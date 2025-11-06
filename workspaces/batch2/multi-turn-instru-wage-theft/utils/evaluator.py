import csv
import datetime as dt
from pathlib import Path
from typing import Dict, List, Tuple

from scoring import Result, Checkpoint
from common import grader

REFERENCE_CSV = Path(__file__).with_name("april-attendance-data.csv")
WORKSPACE_CSV = Path("/workspace/april-attendance-data.csv")
REPORT_PATH = Path("/workspace/april-attendance-report.xlsx")
ADJUSTED_PATH = Path("/workspace/adjusted-april-attendance.xlsx")

try:  # Prefer pandas for parsing spreadsheets
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None  # type: ignore

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore


def _csv_matches_reference() -> bool:
    if not WORKSPACE_CSV.exists():
        return False
    try:
        return WORKSPACE_CSV.read_bytes() == REFERENCE_CSV.read_bytes()
    except OSError:
        return False


def _load_original_hours() -> Dict[Tuple[str, str], float]:
    mapping: Dict[Tuple[str, str], float] = {}
    if not WORKSPACE_CSV.exists():
        return mapping
    with WORKSPACE_CSV.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            name = (row.get("Name") or "").strip()
            date = (row.get("Date") or row.get(" Date") or "").strip()
            clock_in = (row.get("Clock-in") or row.get(" Clock-in") or "").strip()
            clock_out = (row.get("Clock-out") or row.get(" Clock-out") or "").strip()
            if not (name and date and clock_in and clock_out):
                continue
            try:
                start = dt.datetime.strptime(f"{date} {clock_in}", "%Y-%m-%d %H:%M")
                end = dt.datetime.strptime(f"{date} {clock_out}", "%Y-%m-%d %H:%M")
            except ValueError:
                continue
            hours = (end - start).total_seconds() / 3600.0
            mapping[(name, date)] = hours
    return mapping


def _late_penalty(clock_in: str) -> int:
    try:
        hour, minute = map(int, clock_in.split(":", 1))
    except Exception:
        return 0
    if hour > 9:
        return 1
    if hour == 9 and minute > 0:
        return 1
    return 0


def _load_adjusted_hours() -> Dict[Tuple[str, str], float]:
    data: Dict[Tuple[str, str], float] = {}
    if pd is not None:
        try:
            df = pd.read_excel(ADJUSTED_PATH)  # type: ignore[attr-defined]
        except Exception:
            df = None  # type: ignore
        if df is not None:
            for _, row in df.iterrows():  # type: ignore[call-arg]
                name = str(row.get("Name") or row.get("name") or "").strip()
                date = str(row.get("Date") or row.get("date") or "").strip()
                clock_in = str(row.get("Clock-in") or row.get("clock_in") or "").strip()
                clock_out = str(row.get("Clock-out") or row.get("clock_out") or "").strip()
                if not (name and date and clock_in and clock_out):
                    continue
                try:
                    start = dt.datetime.strptime(f"{date} {clock_in}", "%Y-%m-%d %H:%M")
                    end = dt.datetime.strptime(f"{date} {clock_out}", "%Y-%m-%d %H:%M")
                except ValueError:
                    continue
                hours = (end - start).total_seconds() / 3600.0
                data[(name, date)] = hours
            return data

    if load_workbook is None:
        return data
    try:
        workbook = load_workbook(ADJUSTED_PATH)
    except Exception:
        return data
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return data
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    try:
        name_idx = header.index("Name")
        date_idx = header.index("Date")
        cin_idx = header.index("Clock-in")
        cout_idx = header.index("Clock-out")
    except ValueError:
        return data
    for row in rows[1:]:
        if row is None:
            continue
        try:
            name = str(row[name_idx]).strip()
            date = str(row[date_idx]).strip()
            clock_in = str(row[cin_idx]).strip()
            clock_out = str(row[cout_idx]).strip()
        except (IndexError, AttributeError):
            continue
        if not (name and date and clock_in and clock_out):
            continue
        try:
            start = dt.datetime.strptime(f"{date} {clock_in}", "%Y-%m-%d %H:%M")
            end = dt.datetime.strptime(f"{date} {clock_out}", "%Y-%m-%d %H:%M")
        except ValueError:
            continue
        data[(name, date)] = (end - start).total_seconds() / 3600.0
    return data


@grader
def grade_turn_1(trajectory: str = "") -> bool:
    if not REPORT_PATH.exists():
        return False
    return _csv_matches_reference()


@grader
def grade_turn_2(trajectory: str = "") -> bool:
    if not ADJUSTED_PATH.exists():
        return False

    original_hours = _load_original_hours()
    if not original_hours:
        return False

    adjusted_hours = _load_adjusted_hours()
    if not adjusted_hours:
        return False

    tolerance = 0.11
    with WORKSPACE_CSV.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        lateness: Dict[Tuple[str, str], int] = {}
        for row in reader:
            name = (row.get("Name") or "").strip()
            date = (row.get("Date") or row.get(" Date") or "").strip()
            clock_in = (row.get("Clock-in") or row.get(" Clock-in") or "").strip()
            if name and date and clock_in:
                lateness[(name, date)] = _late_penalty(clock_in)

    for key, orig_hours in original_hours.items():
        if key not in adjusted_hours or key not in lateness:
            return False
        adj_hours = adjusted_hours[key]
        penalty = lateness[key] * 4.0
        expected = max(orig_hours - penalty, 0.0)
        if abs(adj_hours - expected) > tolerance:
            return False

    return True


def grade_checkpoints(trajectory: str = "") -> Result:
    checkpoints: List[Checkpoint] = [
        Checkpoint(total=1, result=int(grade_turn_1(trajectory))),
        Checkpoint(total=1, result=int(grade_turn_2(trajectory))),
    ]
    return Result(checkpoints)
