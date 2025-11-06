import csv
import datetime as dt
from pathlib import Path
from typing import List

from scoring import Result, Checkpoint
from common import grader

WORKSPACE_CSV = Path("/workspace/april-attendance-data.csv")
REFERENCE_CSV = Path(__file__).with_name("april-attendance-data.csv")
OUTPUT_XLSX = Path("/workspace/cleaned-attendance.xlsx")

try:  # pandas offers the simplest reader if available
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None  # type: ignore

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore


def _load_hours_from_csv(path: Path) -> List[float]:
    hours: List[float] = []
    if not path.exists():
        return hours
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            date = (row.get("Date") or row.get(" date") or "").strip()
            start = (row.get("Clock-in") or row.get(" Clock-in") or "").strip()
            end = (row.get("Clock-out") or row.get(" Clock-out") or "").strip()
            if not (date and start and end):
                continue
            try:
                begin = dt.datetime.strptime(f"{date} {start}", "%Y-%m-%d %H:%M")
                finish = dt.datetime.strptime(f"{date} {end}", "%Y-%m-%d %H:%M")
            except ValueError:
                continue
            hours.append((finish - begin).total_seconds() / 3600.0)
    return hours


def _load_hours_from_output(path: Path) -> List[float]:
    hours: List[float] = []
    if not path.exists():
        return hours

    if pd is not None:
        try:
            df = pd.read_excel(path)  # type: ignore[attr-defined]
        except Exception:
            df = None  # type: ignore
        if df is not None:
            for _, row in df.iterrows():  # type: ignore[call-arg]
                date = (str(row.get("Date")) or str(row.get("date"))).strip()
                start = (str(row.get("Clock-in")) or str(row.get("clock_in"))).strip()
                end = (str(row.get("Clock-out")) or str(row.get("clock_out"))).strip()
                if not (date and start and end):
                    continue
                try:
                    begin = dt.datetime.strptime(f"{date} {start}", "%Y-%m-%d %H:%M")
                    finish = dt.datetime.strptime(f"{date} {end}", "%Y-%m-%d %H:%M")
                except ValueError:
                    continue
                hours.append((finish - begin).total_seconds() / 3600.0)
            return hours

    if load_workbook is None:
        return hours
    try:
        workbook = load_workbook(path)
    except Exception:
        return hours
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return hours
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    try:
        date_idx = header.index("Date")
        start_idx = header.index("Clock-in")
        end_idx = header.index("Clock-out")
    except ValueError:
        return hours
    for row in rows[1:]:
        if row is None:
            continue
        try:
            date = str(row[date_idx]).strip()
            start = str(row[start_idx]).strip()
            end = str(row[end_idx]).strip()
        except (IndexError, AttributeError):
            continue
        if not (date and start and end):
            continue
        try:
            begin = dt.datetime.strptime(f"{date} {start}", "%Y-%m-%d %H:%M")
            finish = dt.datetime.strptime(f"{date} {end}", "%Y-%m-%d %H:%M")
        except ValueError:
            continue
        hours.append((finish - begin).total_seconds() / 3600.0)
    return hours


@grader
def grade_turn_1(trajectory: str = "") -> bool:
    if not WORKSPACE_CSV.exists():
        return False
    try:
        workspace_bytes = WORKSPACE_CSV.read_bytes()
        reference_bytes = REFERENCE_CSV.read_bytes()
    except OSError:
        return False
    return workspace_bytes == reference_bytes


@grader
def grade_turn_2(trajectory: str = "") -> bool:
    if not OUTPUT_XLSX.exists():
        return False

    input_hours = _load_hours_from_csv(WORKSPACE_CSV)
    output_hours = _load_hours_from_output(OUTPUT_XLSX)
    if not output_hours:
        return False

    if any(hour > 9.5 for hour in output_hours):
        return False

    if any(hour > 9.5 for hour in input_hours) and len(output_hours) >= len(input_hours):
        return False

    return True


def grade_checkpoints(trajectory: str = "") -> Result:
    checkpoints: List[Checkpoint] = [
        Checkpoint(total=1, result=int(grade_turn_1(trajectory))),
        Checkpoint(total=1, result=int(grade_turn_2(trajectory))),
    ]
    return Result(checkpoints)
